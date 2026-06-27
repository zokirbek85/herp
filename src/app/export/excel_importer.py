"""Excel fayldan ma'lumotlarni o'qish (generic import qatlami)."""

from pathlib import Path
from typing import Any

from openpyxl import load_workbook


def read_excel_rows(path: Path, *, header_row: int = 1) -> list[dict[str, Any]]:
    """Birinchi qatorni ustun nomi sifatida o'qiydi, qolganlarini dict ro'yxati qilib qaytaradi.

    Bo'sh qatorlar (barcha ustunlar `None`) e'tiborsiz qoldiriladi.
    """
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active

    rows_iter = sheet.iter_rows(min_row=header_row, values_only=True)
    headers = [str(header).strip() if header is not None else "" for header in next(rows_iter)]

    result: list[dict[str, Any]] = []
    for raw_row in rows_iter:
        if all(value is None for value in raw_row):
            continue
        result.append({headers[i]: raw_row[i] for i in range(len(headers)) if i < len(raw_row)})
    return result
