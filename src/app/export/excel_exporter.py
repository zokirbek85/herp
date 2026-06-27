"""Har qanday jadvalni Excel (.xlsx) formatiga eksport qilish."""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.export.column_spec import ExportColumn

_HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
_HEADER_FONT = Font(color="FFFFFF", bold=True)


def export_to_excel(
    path: Path,
    columns: list[ExportColumn],
    rows: list[dict],
    *,
    sheet_title: str = "Hisobot",
) -> Path:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = sheet_title

    for col_idx, column in enumerate(columns, start=1):
        cell = sheet.cell(row=1, column=col_idx, value=column.header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    for row_idx, row in enumerate(rows, start=2):
        for col_idx, column in enumerate(columns, start=1):
            sheet.cell(row=row_idx, column=col_idx, value=column.format_value(row))

    for col_idx, column in enumerate(columns, start=1):
        widths = [len(column.header)] + [len(column.format_value(row)) for row in rows]
        sheet.column_dimensions[get_column_letter(col_idx)].width = min(max(widths) + 2, 60)

    sheet.freeze_panes = "A2"

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)
    return path
